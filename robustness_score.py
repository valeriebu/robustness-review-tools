#!/usr/bin/env python3
"""
robustness_score.py
───────────────────
Rule-based robustness scoring pipeline for AI mental health RCT codebook data.

Usage:
    python robustness_score.py input.xlsx
    python robustness_score.py input.xlsx --out results.json
    python robustness_score.py input.xlsx --out results.json --pretty
    python robustness_score.py input.xlsx --csv          # also write results.csv
    python robustness_score.py input.xlsx --mismatch "Tool1=2,Tool2=0"

Input format:
    Excel file (.xlsx) with one sheet per tool.
    Each sheet uses codebook format: column A = field name, column B = extracted value.
    A non-empty, non-"None" value in column B = Yes (scored).
    An empty or "None" value = No (not scored).
    For graded questions (D10 features/target/model-fit):
        A filled cell = 1 by default. Override to 0 or 2 via --graded flag or edit JSON.

Output:
    JSON file with one entry per tool containing:
        - tool name (sheet name)
        - raw answers for every question (True/False/int)
        - dimension scores (0.0–10.0 each)
        - summary (mean score, answered question count)

Mismatch penalty (D1):
    Dimension 1 includes a manual mismatch adjuster: for each major mismatch
    identified between intended, training, and RCT populations, subtract 0.5
    (max −2.0). Set per tool via --mismatch flag or edit the JSON output and
    re-score with --from-json.

Dependencies:
    pip install openpyxl
"""

import argparse
import json
import sys
import csv
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# ROW MAPPING
# Maps 0-based Excel row index → list of scoring question IDs
# Row index 0 = first row of the sheet (the blank row in the template)
# ─────────────────────────────────────────────────────────────────────────────

ROW_MAP = {
    # D1 — Intended Use (Sec 2.2)
    15:  ["d1_ri"],   # Region/Location of Intended Use
    17:  ["d1_si"],   # Setting of Intended Use
    25:  ["d1_xi"],   # Sex of Target Population
    29:  ["d1_ei"],   # Race/Ethnicity of Target Population
    31:  ["d1_ai"],   # Age of Target Population
    39:  ["d1_mi"],   # Mental Health of Target Population
    47:  ["d1_ui"],   # Intended End Users
    52:  ["d1_ap"],   # Restrictions of applicability

    # D1 — Training data origin (Sec 6.2)
    703: ["d1_ot"],   # Dataset origin (training)
    704: ["d1_tf"],   # Timeframe of data collection (training)
    706: ["d1_lt"],   # Location/country (training)
    711: ["d1_st"],   # Setting during data collection (training)

    # D1 — Training demographics (Sec 6.3)
    724: ["d1_xt"],   # Sex in training data
    728: ["d1_et"],   # Race/Ethnicity in training data [1pt — double-weighted]
    732: ["d1_at"],   # Age in training data
    748: ["d1_mt"],   # Mental health in training data
    736: ["d1_ss"],   # SES/employment in training data

    # D1 — Cross-study / RCT (Sec 7.1, 4.2, 4.5, 4.3)
    924: ["d1_sm"],              # Similarities/differences study vs training data
    89:  ["d1_bp", "d6_dm"],    # Baseline population chars [also D6 demographics]
    157: ["d1_cr", "d6_or"],    # Study location / country [also D6 origin]
    162: ["d1_sr", "d6_se"],    # Setting [also D6 setting]
    149: ["d1_lg"],              # Language

    # D2 — Data Integrity & Quality (Sec 6.2, 6.4)
    764: ["d2_n"],    # Total N in training dataset
    765: ["d2_cl"],   # Data cleaning steps
    766: ["d2_mv"],   # Missing values handling [2pt — critical]
    767: ["d2_im"],   # Imputation / transformation
    768: ["d2_qc"],   # Other QC steps
    770: ["d2_gt"],   # Ground truth establishment [2pt — critical, triggers cap]
    769: ["d2_sy"],   # Synthetic data assessment

    # D3 — Initial Development (Sec 3.1, 6.1)
    59:  ["d3_at"],   # Type of algorithm
    60:  ["d3_ar"],   # Rationale for algorithm
    62:  ["d3_ac"],   # Architecture described
    699: ["d3_sp"],   # Data splitting [critical — triggers cap]
    700: ["d3_re"],   # Resampling method
    701: ["d3_op"],   # Optimisation method
    702: ["d3_hp"],   # Hyperparameter optimisation

    # D4 — Internal Validation (Sec 6.5)
    772: ["d4_me"],   # Performance metrics [2pt — critical, triggers 0/10]
    778: ["d4_ra"],   # Rationale for metrics
    779: ["d4_ci"],   # CIs for internal validation [2pt]
    780: ["d4_ad"],   # Additional robustness info

    # D5 & D6 — External Validation (Sec 6.6, 6.7, 6.8, 6.10)
    # Row 781 text is PARSED (not binary): "retrospective" → d5_co=True; "prospective" → d6_co=True; "both" → both True
    # Quality items are the same rows for both D5 and D6 — the gate determines which dimension activates
    781: ["d5_co", "d6_co"],  # GATE row — text parsed in parse_sheet()
    788: ["d5_or", "d6_or"],  # Dataset origin (ext validation study)
    792: ["d5_tf", "d6_tf"],  # Timeframe (ext validation)
    796: ["d5_lo", "d6_lo"],  # Location (ext validation)
    800: ["d5_se", "d6_se"],  # Setting (ext validation)
    804: ["d5_au", "d6_au"],  # Autonomy level (ext validation)
    821: ["d5_sx", "d6_sx"],  # Sex (ext validation demographics)
    825: ["d5_ra", "d6_ra"],  # Race/Ethnicity (ext validation) [1pt]
    829: ["d5_ag", "d6_ag"],  # Age (ext validation)
    845: ["d5_mh", "d6_mh"],  # Mental health (ext validation)
    879: ["d5_cl", "d6_cl"],  # Data cleaning (ext validation)
    880: ["d5_mv", "d6_mv"],  # Missing values (ext validation)
    884: ["d5_gt", "d6_gt"],  # Ground truth (ext validation)
    886: ["d5_me", "d6_me"],  # Performance metrics (ext validation) [2pt]
    892: ["d5_mr", "d6_mr"],  # Rationale for metrics (ext validation)
    894: ["d5_ci", "d6_ci"],  # CIs (ext validation)
    893: ["d5_ts", "d6_ts"],  # Testing strategy rationale (ext validation)

    # D7 — Subgroup & Fairness Testing (Sec 7.1)
    912: ["d7_sx"],   # Sexes
    913: ["d7_ra"],   # Race/ethnicities [1pt]
    914: ["d7_ag"],   # Age groups
    918: ["d7_cl"],   # Clinical groups [1pt]
    906: ["d7_lo"],   # Locations/sites
    905: ["d7_ts"],   # Timespans
    921: ["d7_in"],   # Intersectionality
    922: ["d7_co"],   # Consistencies/inconsistencies described
    926: ["d7_fa"],   # Fairness assessment [2pt]
    927: ["d7_er"],   # Prediction errors by subgroup

    # D8 — Deployment & Operational Stability (Sec 3.1, 7.1, 4.5, 4.6, 4.7, 4.8)
    67:  ["d8_th"],   # Decision threshold
    68:  ["d8_un"],   # Uncertainty scores on UX
    69:  ["d8_es"],   # Escalation channel
    70:  ["d8_ad"],   # System adapts to environments
    907: ["d8_se"],   # Tested across settings
    909: ["d8_dv"],   # Tested across devices
    908: ["d8_av"],   # Tested across autonomy levels
    920: ["d8_us"],   # Tested across end-user characteristics
    163: ["d8_ar"],   # Autonomy level during RCT
    164: ["d8_dr"],   # Device used during study
    167: ["d8_ti"],   # Technical integration during study
    168: ["d8_ut"],   # End-user type during study
    169: ["d8_ue"],   # End-user expertise/training
    673: ["d8_bl"],   # Outcome blinding / bias prevention

    # D9 — Stress-Testing & Sensitivity (Sec 7.1, 4.9)
    923: ["d9_cs"],   # Cross-study comparisons
    925: ["d9_tr"],   # Different training strategies
    928: ["d9_hp"],   # Hyperparameter sensitivity
    929: ["d9_ms"],   # Data missingness impact
    930: ["d9_ba"],   # Data balancing effects
    931: ["d9_ot"],   # Other robustness methods
    679: ["d9_pl"],   # Performance limitations reported
    678: ["d9_rc"],   # Robustness claims with justification
    680: ["d9_ae"],   # Adverse events recorded

    # D10 — Construct Validity (Sec 3.1, 5.1)
    66:  ["d10_or"],  # Output relevance justified
    # Prediction model construct validity (Sec 5.1)
    682: ["d10_fw"],  # Theoretical framework (model)
    683: ["d10_fe"],  # Features appropriate — GRADED 0/1/2 (model)
    684: ["d10_tg"],  # Target appropriate — GRADED 0/1/2 (model)
    685: ["d10_mf"],  # Model-task fit — GRADED 0/1/2 (model)
    686: ["d10_ex"],  # External science cited (model)
    687: ["d10_ep"],  # Explainability for construct claims (model)
    # Chatbot construct validity (Sec 5.1 — chatbot variant)
    # Parallel questions for chatbot tools — same D10 IDs, different rows
    689: ["d10_fw"],  # Theoretical framework (chatbot)
    690: ["d10_fe"],  # Features appropriate — chatbot asks right inputs (GRADED)
    691: ["d10_tg"],  # Target appropriate — chatbot responses appropriate (GRADED)
    692: ["d10_mf"],  # Model-task fit — chatbot connects inputs to outputs (GRADED)
    693: ["d10_ex"],  # External science cited (chatbot)
    694: ["d10_ep"],  # Explainability used (chatbot)
}

# Questions that use graded scoring (0 = not mentioned, 1 = claimed, 2 = evidenced)
GRADED_QUESTIONS = {"d10_fe", "d10_tg", "d10_mf"}


# ─────────────────────────────────────────────────────────────────────────────
# DIMENSION CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

DIMENSIONS = {
    1: {
        "label": "Data Representativeness & Scope Match",
        "formula": "(raw / 13.5) × 10 − (mismatches × 0.5)",
        "scoring": "compare",
        "denom": 13.5,
        "cap_intended": 5.0,      # if no intended use specified → cap at 5/10
        "mismatch_id": "d1_mm",   # manual adjuster (0–4)
        "intended_qs": ["d1_ri","d1_si","d1_xi","d1_ei","d1_ai","d1_mi","d1_ui","d1_ap"],
        "questions": {
            "d1_ri": 0.5, "d1_si": 0.5, "d1_xi": 0.5, "d1_ei": 0.5,
            "d1_ai": 0.5, "d1_mi": 0.5, "d1_ui": 0.5, "d1_ap": 1.0,
            "d1_ot": 1.0, "d1_tf": 0.5, "d1_lt": 0.5, "d1_st": 0.5,
            "d1_xt": 0.5, "d1_et": 1.0, "d1_at": 0.5, "d1_mt": 0.5,
            "d1_ss": 0.5, "d1_sm": 1.0, "d1_bp": 1.0,
            "d1_cr": 0.5, "d1_sr": 0.5, "d1_lg": 0.5,
        },
    },
    2: {
        "label": "Data Integrity & Quality",
        "formula": "(raw / 8) × 10",
        "scoring": "weight",
        "denom": 8,
        "cap_if_absent": ("d2_gt", 4.0),   # ground truth absent → cap 4/10
        "questions": {
            "d2_n": 1.0, "d2_cl": 1.0, "d2_mv": 2.0, "d2_im": 1.0,
            "d2_qc": 1.0, "d2_gt": 2.0, "d2_sy": 1.0,
        },
    },
    3: {
        "label": "Initial Development",
        "formula": "(raw / 7) × 10",
        "scoring": "weight",
        "denom": 7,
        "cap_if_absent": ("d3_sp", 4.0),   # no data splitting → cap 4/10
        "questions": {
            "d3_at": 1.0, "d3_ar": 1.0, "d3_ac": 1.0, "d3_sp": 1.0,
            "d3_re": 1.0, "d3_op": 1.0, "d3_hp": 1.0,
        },
    },
    4: {
        "label": "Internal Validation",
        "formula": "(raw / 6) × 10",
        "scoring": "weight",
        "denom": 6,
        "zero_if_absent": "d4_me",         # no metrics → score = 0
        "questions": {
            "d4_me": 2.0, "d4_ra": 1.0, "d4_ci": 2.0, "d4_ad": 1.0,
        },
    },
    5: {
        "label": "Retrospective External Validation",
        "formula": "GATE → if absent: 0/10; if present: (quality / 6) × 10",
        "scoring": "gate",
        "gate_id": "d5_co",
        "gate_type": "retrospective",  # gate row 781 parsed for 'retrospective'/'both'
        "denom": 6,
        "questions": {
            "d5_or": 1.0, "d5_tf": 0.5, "d5_lo": 0.5, "d5_se": 0.5,
            "d5_au": 0.5, "d5_sx": 0.5, "d5_ra": 1.0, "d5_ag": 0.5,
            "d5_mh": 0.5, "d5_cl": 0.5, "d5_mv": 0.5, "d5_gt": 0.5,
            "d5_me": 2.0, "d5_mr": 0.5, "d5_ci": 1.0, "d5_ts": 0.5,
        },
    },
    6: {
        "label": "Prospective External Validation",
        "formula": "GATE → if absent: 0/10; if present: (quality / 6) × 10",
        "scoring": "gate",
        "gate_id": "d6_co",
        "gate_type": "prospective",  # gate row 781 parsed for 'prospective'/'both'
        "denom": 6,
        "questions": {
            "d6_or": 1.0, "d6_tf": 0.5, "d6_lo": 0.5, "d6_se": 0.5,
            "d6_au": 0.5, "d6_sx": 0.5, "d6_ra": 1.0, "d6_ag": 0.5,
            "d6_mh": 0.5, "d6_cl": 0.5, "d6_mv": 0.5, "d6_gt": 0.5,
            "d6_me": 2.0, "d6_mr": 0.5, "d6_ci": 1.0, "d6_ts": 0.5,
        },
    },
    7: {
        "label": "Subgroup & Fairness Testing",
        "formula": "(raw / 10) × 10",
        "scoring": "weight",
        "denom": 10,
        "questions": {
            "d7_sx": 0.5, "d7_ra": 1.0, "d7_ag": 0.5, "d7_cl": 1.0,
            "d7_lo": 0.5, "d7_ts": 0.5, "d7_in": 1.0, "d7_co": 1.0,
            "d7_fa": 2.0, "d7_er": 1.0,
        },
    },
    8: {
        "label": "Deployment & Operational Stability",
        "formula": "(raw / 10) × 10",
        "scoring": "weight",
        "denom": 10,
        "questions": {
            "d8_th": 1.0, "d8_un": 1.0, "d8_es": 1.0, "d8_ad": 1.0,
            "d8_se": 0.5, "d8_dv": 0.5, "d8_av": 0.5, "d8_us": 0.5,
            "d8_ar": 1.0, "d8_dr": 0.5, "d8_ti": 1.0,
            "d8_ut": 0.5, "d8_ue": 1.0, "d8_bl": 1.0,
        },
    },
    9: {
        "label": "Stress-Testing & Sensitivity",
        "formula": "(raw / 9) × 10",
        "scoring": "weight",
        "denom": 9,
        "questions": {
            "d9_cs": 1.0, "d9_tr": 1.0, "d9_hp": 1.0, "d9_ms": 1.0,
            "d9_ba": 1.0, "d9_ot": 1.0, "d9_pl": 1.0, "d9_rc": 1.0,
            "d9_ae": 1.0,
        },
    },
    10: {
        "label": "Construct Validity",
        "formula": "(raw / 11) × 10",
        "scoring": "weight",
        "denom": 11,
        "questions": {
            "d10_or": 1.0, "d10_fw": 1.0,
            "d10_fe": 2.0,  # graded
            "d10_tg": 2.0,  # graded
            "d10_mf": 2.0,  # graded
            "d10_ex": 2.0,
            "d10_ep": 1.0,
        },
    },
}


# ─────────────────────────────────────────────────────────────────────────────
# PARSING
# ─────────────────────────────────────────────────────────────────────────────

def is_filled(val):
    """Return True if a cell value counts as 'Yes' (answered)."""
    if val is None:
        return False
    s = str(val).strip()
    return s != "" and s.lower() != "none"


def parse_sheet(rows, mismatch_count=0):
    """
    Parse a list of Excel rows (from sheet_to_json header=1 or iter_rows)
    into a flat answers dict keyed by question ID.

    Parameters
    ----------
    rows : list of tuples
        Each tuple is (col_A_value, col_B_value, ...).
    mismatch_count : int
        Manual mismatch penalty for D1 (0–4). Set externally.

    Returns
    -------
    dict
        {question_id: True/False/int/None}
    """
    # Initialise all question IDs to None
    all_qids = {qid for qids in ROW_MAP.values() for qid in qids}
    for dim_cfg in DIMENSIONS.values():
        all_qids.update(dim_cfg["questions"].keys())
    answers = {qid: None for qid in all_qids}
    answers["d1_mm"] = max(0, min(4, int(mismatch_count)))  # mismatch adjuster
    answers["_tool_type"] = "other"  # overridden by parse_sheet

    for row_idx, row in enumerate(rows):
        if row_idx not in ROW_MAP:
            continue
        val = row[1] if len(row) > 1 else None
        yes = is_filled(val)
        val_lower = str(val).lower() if val else ""

        for qid in ROW_MAP[row_idx]:
            # Row 781 is the external validation gate — parse text content
            # rather than treating it as a simple yes/no
            if row_idx == 781:
                if qid == "d5_co":
                    answers[qid] = ("retrospective" in val_lower or "both" in val_lower)
                elif qid == "d6_co":
                    answers[qid] = ("prospective" in val_lower or "both" in val_lower)
            elif qid in GRADED_QUESTIONS:
                # graded: filled = 1 (claimed), empty = 0 (not mentioned)
                # user must override to 2 (evidenced) manually in JSON
                answers[qid] = 1 if yes else 0
            else:
                answers[qid] = yes

    # Detect tool type from which construct validity section is filled
    # Row 689 = chatbot framework question; Row 682 = prediction model framework
    is_chatbot = answers.get("d10_fw") and         (rows[689][1] if len(rows) > 689 else None) is not None and         str(rows[689][1] if len(rows) > 689 else "").strip().lower() not in ("","none")
    is_model = (rows[682][1] if len(rows) > 682 else None) is not None and         str(rows[682][1] if len(rows) > 682 else "").strip().lower() not in ("","none")
    answers["_tool_type"] = "chatbot" if is_chatbot else ("model" if is_model else "other")
    return answers


# ─────────────────────────────────────────────────────────────────────────────
# SCORING ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def raw_points(dim_cfg, answers):
    """Sum raw points for a dimension's questions."""
    total = 0.0
    for qid, pts in dim_cfg["questions"].items():
        v = answers.get(qid)
        if qid in GRADED_QUESTIONS:
            total += (v or 0) * (pts / 2)   # 0→0, 1→half pts, 2→full pts
        elif v is True:
            total += pts
    return total


def score_dimension(dim_id, dim_cfg, answers):
    """
    Calculate the 0–10 score for one dimension.
    Returns (score_float, raw_float, note_str).
    """
    tool_type = answers.get("_tool_type", "other")

    # ── Chatbot-specific overrides ────────────────────────────────────────────
    if tool_type == "chatbot":
        if dim_id == 2:
            # D2 (Data Integrity & Quality) not applicable for chatbots —
            # assumes ML training dataset; chatbots have no such dataset
            return None, 0.0, "N/A — chatbot: no ML training dataset"
        if dim_id == 3:
            # D3 chatbot variant: only algorithm type, rationale, architecture
            # Data splitting / resampling / optimisation / hyperparams not applicable
            chatbot_qs = {"d3_at": 1.0, "d3_ar": 1.0, "d3_ac": 1.0}
            raw = sum(pts for qid, pts in chatbot_qs.items()
                      if answers.get(qid) is True)
            s = round((raw / 3.0) * 10, 1)
            note = "Chatbot variant: only algorithm type, rationale, architecture scored (÷3)"
            return max(0.0, min(10.0, s)), round(raw, 2), note
    # ─────────────────────────────────────────────────────────────────────────
    scoring = dim_cfg["scoring"]

    if scoring == "gate":
        gate_val = answers.get(dim_cfg["gate_id"])
        if gate_val is not True:
            return 0.0, 0.0, "Gate = No — dimension scores 0"
        raw = raw_points(dim_cfg, answers)
        s = round(min(10.0, (raw / dim_cfg["denom"]) * 10), 1)
        return s, raw, "Gate = Yes — quality items scored"

    raw = raw_points(dim_cfg, answers)
    s = (raw / dim_cfg["denom"]) * 10
    note = ""

    if scoring == "compare":
        mm = answers.get(dim_cfg["mismatch_id"], 0) or 0
        s -= mm * 0.5
        if mm:
            note += f"Mismatch penalty applied: {mm} × −0.5 = −{mm*0.5}. "
        intended = [answers.get(q) for q in dim_cfg["intended_qs"]]
        if not any(intended):
            s = min(s, dim_cfg["cap_intended"])
            note += f"Intended use unspecified → capped at {dim_cfg['cap_intended']}/10. "

    if "cap_if_absent" in dim_cfg:
        crit_q, cap_val = dim_cfg["cap_if_absent"]
        if not answers.get(crit_q):
            s = min(s, cap_val)
            note += f"Critical item '{crit_q}' absent → capped at {cap_val}/10. "

    if "zero_if_absent" in dim_cfg:
        if not answers.get(dim_cfg["zero_if_absent"]):
            s = 0.0
            note += f"Critical item '{dim_cfg['zero_if_absent']}' absent → score = 0. "

    s = round(max(0.0, min(10.0, s)), 1)
    return s, round(raw, 2), note.strip()


def score_tool(answers):
    """
    Score all 10 dimensions for one tool.

    Returns
    -------
    dict with keys 'dimensions' (list of per-dim results) and 'summary'.
    """
    results = []
    for dim_id in range(1, 11):
        cfg = DIMENSIONS[dim_id]
        s, raw, note = score_dimension(dim_id, cfg, answers)
        yes_count = sum(
            1 for q in cfg["questions"]
            if (q in GRADED_QUESTIONS and (answers.get(q) or 0) > 0)
            or (q not in GRADED_QUESTIONS and answers.get(q) is True)
        )
        results.append({
            "dimension": dim_id,
            "label": cfg["label"],
            "score": s,
            "raw_points": raw,
            "questions_answered_yes": yes_count,
            "questions_total": len(cfg["questions"]),
            "formula": cfg["formula"],
            "note": note,
        })

    valid_scores = [r["score"] for r in results if r["score"] is not None and r["score"] >= 0]
    mean = round(sum(valid_scores) / len(valid_scores), 2) if valid_scores else None
    return {
        "dimensions": results,
        "summary": {
            "mean_score": mean,
            "mismatch_adjuster_d1": answers.get("d1_mm", 0),
            "graded_question_defaults": {
                "note": "d10_fe/d10_tg/d10_mf default to 1 (claimed without evidence) if filled. Edit these in the JSON output and re-run with --from-json to apply 0 or 2.",
                "d10_fe": answers.get("d10_fe"),
                "d10_tg": answers.get("d10_tg"),
                "d10_mf": answers.get("d10_mf"),
            }
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL → JSON PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def process_excel(xlsx_path, mismatch_map=None):
    """
    Read all sheets from an Excel file, parse and score each one.

    Parameters
    ----------
    xlsx_path : str or Path
    mismatch_map : dict, optional
        {sheet_name: mismatch_count} — manual D1 adjuster per tool.

    Returns
    -------
    list of dicts, one per sheet.
    """
    mismatch_map = mismatch_map or {}
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    output = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        mm = mismatch_map.get(sheet_name, 0)
        answers = parse_sheet(rows, mismatch_count=mm)
        scored = score_tool(answers)

        output.append({
            "tool_name": sheet_name,
            "source_sheet": sheet_name,
            "answers": answers,
            **scored,
        })
        print(f"  ✓ {sheet_name}: mean={scored['summary']['mean_score']}")

    wb.close()
    return output


# ─────────────────────────────────────────────────────────────────────────────
# CSV EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def write_csv(results, csv_path):
    """Write a summary CSV — one row per tool, one column per dimension score."""
    fieldnames = ["tool_name", "mean_score"] + [f"D{i}" for i in range(1, 11)]
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in results:
            row = {"tool_name": r["tool_name"],
                   "mean_score": r["summary"]["mean_score"]}
            for dim in r["dimensions"]:
                row[f"D{dim['dimension']}"] = dim["score"]
            writer.writerow(row)
    print(f"  ✓ CSV written to {csv_path}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def parse_mismatch_arg(s):
    """Parse 'Tool1=2,Tool2=0' into {'Tool1': 2, 'Tool2': 0}."""
    result = {}
    if not s:
        return result
    for part in s.split(","):
        part = part.strip()
        if "=" in part:
            name, val = part.split("=", 1)
            try:
                result[name.strip()] = int(val.strip())
            except ValueError:
                pass
    return result


def main():
    parser = argparse.ArgumentParser(
        description="Score AI robustness from codebook Excel files.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("input", help="Input .xlsx file")
    parser.add_argument("--out", default=None,
                        help="Output JSON path (default: <input>_scored.json)")
    parser.add_argument("--pretty", action="store_true",
                        help="Pretty-print JSON output (indent=2)")
    parser.add_argument("--csv", action="store_true",
                        help="Also write a summary CSV file")
    parser.add_argument("--mismatch", default="",
                        help="D1 mismatch penalty per tool, e.g. 'Tool1=2,Tool2=1'")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: File not found: {input_path}")
        sys.exit(1)
    if input_path.suffix.lower() != ".xlsx":
        print("ERROR: Input must be a .xlsx file")
        sys.exit(1)

    out_path = Path(args.out) if args.out else Path(input_path.stem + "_scored.json")
    mismatch_map = parse_mismatch_arg(args.mismatch)

    print(f"\nProcessing: {input_path}")
    results = process_excel(input_path, mismatch_map)

    # Write JSON
    json_str = json.dumps(results, indent=2 if args.pretty else None, ensure_ascii=False)
    out_path.write_text(json_str, encoding="utf-8")
    print(f"  ✓ JSON written to {out_path}")

    # Optional CSV
    if args.csv:
        csv_path = out_path.with_suffix(".csv")
        write_csv(results, csv_path)

    # Print summary table
    print(f"\n{'Tool':<30} {'D1':>4} {'D2':>4} {'D3':>4} {'D4':>4} {'D5':>4} "
          f"{'D6':>4} {'D7':>4} {'D8':>4} {'D9':>4} {'D10':>4} {'Mean':>5}")
    print("-" * 85)
    for r in results:
        scores = [d["score"] for d in r["dimensions"]]
        row = f"{r['tool_name']:<30}"
        for s in scores:
            row += f" {str(s) if s is not None else "N/A":>4}"
        row += f" {r['summary']['mean_score']:>5}"
        print(row)

    print(f"\nDone. {len(results)} tool(s) scored.\n")


if __name__ == "__main__":
    main()
